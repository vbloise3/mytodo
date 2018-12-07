import datetime
import boto3
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

def createworkbook():
    BUCKET = 'vbloise3acloudgurus3website'
    session = boto3.Session(aws_access_key_id='',
                            aws_secret_access_key='')
    s3 = session.resource('s3')
    wb = Workbook()
    ws = wb.active
    wb.active.title = "openpyxl sheet 1"
    ws['A1'] = 42
    ws['A2'] = datetime.datetime.now()
    ws['A2'].style = NamedStyle(name='datetime', number_format='DD/MM/YY')
    ws.append([1,2,3])
    wb.save('openpyxlFirst.xlsx')
    s3.meta.client.upload_file('openpyxlFirst.xlsx', BUCKET, 'openpyxlFirst.xlsx')
    print("workbook saved")

createworkbook()